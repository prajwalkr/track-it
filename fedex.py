import urllib
from bs4 import BeautifulSoup
import requests
import openpyxl as xl
from dateutil.parser import parse
from datetime import datetime

__author__ = 'K R Prajwal'

'''
	If an exception is raised anywhere, try once with an updated cookie in line no 101
'''

class Fedex_Transit_Times(object):
	'''
		This class reads an Excel spreadsheet and adds the following columns:

			Column 2 - FedEx Economy transit time
			Column 3 - FedEx Priority Overnight transit time (optional)
			Column 4 - FedEx Standard Overnight transit time (optional)
			Column 5 - Number of days for Economy
			Column 6 - Number of days for Priority Overnight
			Column 7 - Number of days for Standard Overnight

		for each Destination pincode row in the spreadsheet
	'''

	def __init__(self,excel_file):
		'''
			Class attributes:
				source_file - path of excel file where the data is read and written
				wb - Current active workbook
				std_overnight
				prioirty_overnight
				economy
				From pincode (default is 400057)
		'''

		self.source_file = excel_file
		try:
			self.wb = xl.load_workbook(excel_file)
		except:
			raise Exception('The excel file is absent')
		self.std_overnight = None
		self.economy = None
		self.prioirty_overnight = None
		self.From = '400057'

	def store_transit_times(self,sheet,row):
		'''
			Stores the transit times and also calls store_differences to 
			store the expected number of transit days.
		'''
		sheet.cell(row=row, column=2).value = self.economy
		sheet = self.store_differences(sheet,row,5,self.economy)

		if self.std_overnight is not None:
			sheet.cell(row=row, column=4).value = self.std_overnight
			sheet = self.store_differences(sheet,row,7,self.std_overnight)

		elif self.prioirty_overnight is not None:
			sheet.cell(row=row, column=3).value = self.prioirty_overnight
			sheet = self.store_differences(sheet,row,6,self.prioirty_overnight)

		return sheet

	def store_differences(self,sheet,row,col,Time):
		# Remove the word 'by' in the Time string for converting to datetime object
		Time = parse(Time[:Time.find('by')] + Time[Time.find('by') + 3:])
		now = datetime.now()
		difference = Time - now
		days = str(difference.days)
		sheet.cell(row=row,column=col).value = days

		return sheet

	def fetch_transit_times(self):
		'''
			Function that is called to fill the excel file with transit times
		'''

		# Assumed only one sheet exists, with name = Sheet1
		sheet = self.wb.get_sheet_by_name('Sheet1')
		col = sheet.columns[0] 			# first column
		row = 1							# row number
		for cell in col:				# for every cell in column 1
			'''
				A try and except block catches any Exception and saves the results before exiting.
			'''
			try:	
				'''
					This if condition ensures that already computed values are not computed again
				'''					
				if sheet.cell(row=row, column=2).value is not None:
					row += 1
					continue

				print row
				self.transit_time(cell.value)

				sheet = self.store_transit_times(sheet,row)
				
			except:
				# save the results and then raise
				self.wb.save(self.source_file)
				raise

			row += 1

		# finally save the results
		self.wb.save(self.source_file)

	def transit_time(self,To):
		'''
			Returns the two transit times for shipment from 'self.From' to 'To'
		'''

		with requests.Session() as s:
			'''
				Starts a session properly with page cookies
				The get request below is, thus, essential
			'''
			url = 'https://www.fedex.com/ratefinder/home?source=gh&cc=in&language=en'
			headers = { 
						'Host': 'www.fedex.com',
						'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:42.0) Gecko/20100101 Firefox/42.0',
						'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
						'Accept-Language': 'en-US,en;q=0.5',
						'Accept-Encoding': 'gzip, deflate',
						'DNT': '1',
						'Cookie': 'siteDC=edc; fdx_cbid=30711172601449727021040500250571; fdx_locale=en_IN; WGRTSESSIONID=4T6KdtH2w5RyfToU6KMqCFZjNOw6qJykGEcO7sbjH27j-9aj08yj!-24471921; mbox=session#1449727026629-291834#1449729312|check#true#1449727512; AMCV_1E22171B520E93BF0A490D44%40AdobeOrg=1999109931%7CMCIDTS%7C16780%7CMCMID%7C67921623005500323240544017274262181973%7CMCAID%7CNONE; s_sess=%20SC_LINKS%3D%3B%20setLink%3Din%252Fen%252Fwgrt%252Fshipfromto%255E%255EWGRT%253AgetTransitTimeBtn-StandardRate%255E%255Ein%252Fen%252Fwgrt%252Fshipfromto%2520%257C%2520WGRT%253AgetTransitTimeBtn-StandardRate%255E%255E%3B; s_pers=%20s_vnum%3D1449772200762%2526vn%253D1%7C1449772200762%3B%20s_visit%3D1%7C1449729509538%3B%20gpv_pageName%3Din%252Fen%252Fwgrt%252Fshipfromto%7C1449729509539%3B%20s_nr%3D1449727709542-New%7C1481263709542%3B%20s_evar54%3D0%7C1449814109544%3B%20s_invisit%3Dtrue%7C1449729509545%3B%20s_prevChan%3D%257C1449727709546%7C1450937309547%3B; s_cc=true; s_sq=fedexglblDev%3D%2526pid%253Dus%25252Fen%25252Ffedex%25252Fglobal%25252Fhomepage%25252Fglobal%252520home%252520page%2526pidt%253D1%2526oid%253Dhttp%25253A%25252F%25252Fwww.fedex.com%25252Fin%25252F%2526ot%253DA; cc=/in/; countryPath=ratefinder; tracking_locale=en_IN; SMIDENTITY=SQocwyCynBZ/xVDsNHPzimuvN42GCh2ttSEv+cS/8CSeZIzRIDTJDOH32B1vK0K2ZKWNk2ww56Yd95YU3Jt/LoOz/LTqtKD8IAArW72ihyOr2Wv6U8FIP05UrSbi87BbpqM2nus5+KgZDoSUw/7mlOizdG9pvKlIGOyIzb0l4fgZi500mDFTm0wxc9idr05FxDrYNHvXSMt5w6pj9fCDLHALtQ7zXcXqi6Mq1Y+uAneizahvBgjTX/or/1K399A6k/UqovyyYxS2fWrrONihOdLG1WYD3IXmKo8ty4kGisdVcno46Br6MtVsw2MMEOfyMKmMWnyLDXbqHm9Rl6EYtIC/U1tNC/I735Z0bNouFeLv4tXmnhRDNJT+6UhhbbZkCt8ORPtRg90NbDtzmOi+Q2NzUloEGnr69oj1LCFOkUe/N2pBt76G/oPXWQhw4TpC3hMChv0Kot+rZcVjJ/2LaKMFfTdLqU9HQzjq9fSMlCsUSUlOKt3MvJz2pd1ZlKamIIwZouV2YLzdz7hTIRmPGak/erW3cXmC',
						'Connection': 'keep-alive'
					  }

			params = {
						'source':'gh',
						'cc':'in',
						'language':'en'
					 }
			response = s.get(url,headers=headers,params=urllib.urlencode(params))

			'''
				Fill the pin codes form below using a post request
			'''

			url = 'https://www.fedex.com/ratefinder/standalone?method=goToPackageInfoPage'
			values = {
					  'BuildTimeStamp':'2015-09-09+AT+19:49:40',
					  'transitTime':'true',
					  'doEdt':'false',
					  'locId':'',
					  'originSelected':'N',
					  'destSelected':'N',
					  'origState':'',
					  'pricingOptionDisplayed':'false',
					  'cmdcResponse': '{"transactionDetails":[{"sourceSystem":"VACS","transactionId":"getAllServicesAndPackaging:vacs-getAllServicesAndPackagingLogic-pje85402-20151209-23:48:16,355-1398570245"}],"output":{"serviceOptions":[{"key":"PRIORITY_OVERNIGHT","displayText":"FedEx+Priority+Overnight<SUP>&reg;</SUP>"},{"key":"STANDARD_OVERNIGHT","displayText":"FedEx+Standard+Overnight<SUP>&reg;</SUP>"},{"key":"FEDEX_EXPRESS_SAVER","displayText":"FedEx+Economy<SUP>&reg;</SUP>"}],"packageOptions":[{"packageType":{"key":"YOUR_PACKAGING","displayText":"Your+Packaging"},"rateTypes":["WEIGHT_BASED"],"subpackageInfoList":[{"description":"Please+enter+the+weight+and+dimensions+of+your+package+for+a+more+accurate+estimated+rate.","dimensionText":""}],"maxMetricWeightAllowed":{"units":"KG","value":68},"maxWeightAllowed":{"units":"LB","value":150}}],"oneRate":false,"pricingOptions":[{"key":"WEIGHT_BASED","displayText":"FedEx+Standard+Rate"},{"key":"FLAT_BASED","displayText":"FedEx+One+Rate"}]},"successful":true}',			  
					  'zipField':'',
					  'currentPage':'rfsshipfromto',
					  'outlookAddressType':'',
					  'outLookResult':'',
					  'origCountry':'IN',
					  'origZip':str(self.From),
					  'origCity':'',
					  'destCountry':'IN',
					  'destZip':str(To),
					  'destCity':'',
					  'pricingOption':"FEDEX_STANDARD_RATE",
					  'totalNumberOfPackages':'1',
					  'isPackageIdentical':'NO',
					  'perPackageWeight':'0.5',
					  'weightUnit':'kgs',
					  'receivedAtCode':'1',
					  'shipDate':'12/10/2015',
					  'shipCalendarDate':'12/10/2015'
					}

			headers['Referer'] = 'https://www.fedex.com/ratefinder/home?source=gh&cc=in&language=en'
			response = s.post(url,data=values,headers=headers)


			''' 
				Now encode the url and data for the second form and make 
				a post request in the same session
			''' 

			url = 'https://www.fedex.com/ratefinder/standalone?method=goToResultSummaryPage'
			values = { 
					 'cc':'IN',
					 'origCountry':'IN',
					 'destCountry':'IN',
					 'language':'en',
					 'locId':'',
					 'submitAction':'',
					 'shipWithAcct':'false',
					 'receivedAtCode':'1',
					 'dropOffPackage':'false',
					 'shipToResidence':'false',
					 'companyType':'Express',
					 'packageCount':'1',
					 'weightUnit':'kgs',
					 'pickupRequestType':'',
					 'displayDimProfile':'false',
					 'loginCountry':'US',
					 'displayFR_InchesDimpProfile':'false',
					 'displayFR_CmDimProfile':'false',
					 'currentPage':'rfspackageinfo',
					 'cmdcResponse':'',
					 'pricingOptionDisplayed':'false',
					 'pricingOption':'FEDEX_STANDARD_RATE',
					 'packageForm.packageList[0].selectedPackageType':'1',
					 'shipContent':'products',
					 'shipmentPurpose':'NOT_SOLD',
					 'customsValue':'999.00',
					 'customsCurrencyUnit':'INR',
					 'freightOnValue':'OWN_RISK',
					 'packageForm.packageList[0].qtyPerProfile':'1',
					 'packageForm.packageList[0].weight':'0.5',
					 'packageForm.packageList[0].weightUnit':'kgs',
					 'packageForm.packageList[0].packageType':'1',
					 'packageForm.packageList[0].dimLength':'L',
					 'packageForm.packageList[0].dimWidth':'W',
					 'packageForm.packageList[0].dimHeight':'H',
					 'packageForm.packageList[0].dimUnit':'cm',
					 'packageForm.packageList[0].declaredValue':'0.00',
					 'packageForm.packageList[0].currencyCode':'INR',
					 'totalDecValue':'0.00',
					 'totalPackageQty':'1',
					 'perPackageWeight':'50',
					 'isPackageIdentical':'NO'
					 }

			# only the Referer key has changed so just change that!
			headers['Referer'] = 'https://www.fedex.com/ratefinder/standalone?method=goToPackageInfoPage'

			response = s.post(url,data=values,headers=headers)

			# extract the three fields
			soup = BeautifulSoup(response.content,'html.parser')

			if soup.find('font',{'id':'STANDARD_OVERNIGHT_dateTime0'}) is not None:
				self.std_overnight = soup.find('font',{'id':'STANDARD_OVERNIGHT_dateTime0'}).string.strip()
			if soup.find('font',{'id':'PRIORITY_OVERNIGHT_dateTime0'}) is not None:
				self.prioirty_overnight = soup.find('font',{'id':'PRIORITY_OVERNIGHT_dateTime0'}).string.strip()
			self.economy = soup.find('font',{'id':'FEDEX_EXPRESS_SAVER_dateTime1'}).string.strip()


f = Fedex_Transit_Times('sample.xlsx')
f.fetch_transit_times()